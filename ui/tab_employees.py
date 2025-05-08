from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QLabel, QComboBox, QCalendarWidget, QTextBrowser,
    QMessageBox, QDialog, QFormLayout, QDialogButtonBox, QTimeEdit,
    QDoubleSpinBox, QPushButton, QLineEdit
)
from PyQt6.QtCore import QDate, QTime
from database.db import Database
from datetime import timedelta, time, datetime
import tempfile, os, subprocess
from openpyxl import Workbook
from openpyxl.styles import Font
import platform


def to_qtime(value):
    if isinstance(value, QTime):
        return value
    elif isinstance(value, str):
        return QTime.fromString(value, "HH:mm:ss")
    elif isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) % 60
        seconds = total_seconds % 60
        return QTime(hours, minutes, seconds)
    elif isinstance(value, time):
        return QTime(value.hour, value.minute, value.second)
    else:
        return QTime()


class AddEmployeeDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Добавить сотрудника")

        self.first_name_edit = QLineEdit()
        self.last_name_edit = QLineEdit()

        layout = QFormLayout()
        layout.addRow("Имя:", self.first_name_edit)
        layout.addRow("Фамилия:", self.last_name_edit)

        self.buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)
        layout.addRow(self.buttons)
        self.setLayout(layout)

    def get_data(self):
        return self.first_name_edit.text(), self.last_name_edit.text()


class EditShiftDialog(QDialog):
    def __init__(self, parent=None, shift_data=None, employees=None):
        super().__init__(parent)
        self.setWindowTitle("Изменить смену")

        self.employee_selector = QComboBox()
        self.start_time_edit = QTimeEdit()
        self.end_time_edit = QTimeEdit()
        self.hourly_rate_edit = QDoubleSpinBox()
        self.hourly_rate_edit.setRange(0, 10000)
        self.hourly_rate_edit.setValue(200.0)
        self.hourly_rate_edit.setSuffix(" ₽/ч")

        layout = QFormLayout()
        layout.addRow("Сотрудник:", self.employee_selector)
        layout.addRow("Начало смены:", self.start_time_edit)
        layout.addRow("Конец смены:", self.end_time_edit)
        layout.addRow("Зарплата в час:", self.hourly_rate_edit)

        self.buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.buttons.accepted.connect(self.accept)
        self.buttons.rejected.connect(self.reject)
        layout.addRow(self.buttons)
        self.setLayout(layout)

        for emp_id, full_name in employees.items():
            self.employee_selector.addItem(full_name, emp_id)

        if shift_data:
            self.employee_selector.setCurrentIndex(
                self.employee_selector.findData(shift_data['employee_id'])
            )
            self.start_time_edit.setTime(to_qtime(shift_data['shift_start']))
            self.end_time_edit.setTime(to_qtime(shift_data['shift_end']))
            hours = self.calculate_hours(shift_data)
            if hours > 0:
                self.hourly_rate_edit.setValue(float(shift_data['shift_salary']) / hours)

        self.shift_data = shift_data

    def get_data(self):
        return (
            self.employee_selector.currentData(),
            self.start_time_edit.time().toString("HH:mm:ss"),
            self.end_time_edit.time().toString("HH:mm:ss"),
            self.hourly_rate_edit.value()
        )

    def calculate_hours(self, shift):
        start = to_qtime(shift['shift_start'])
        end = to_qtime(shift['shift_end'])
        return start.secsTo(end) / 3600


class SummaryDialog(QDialog):
    def __init__(self, parent=None, db=None, employees=None):
        super().__init__(parent)
        self.setWindowTitle("Сводка по сменам")
        self.db = db
        self.employees = employees
        self.summary_data = []

        self.start_date_edit = QCalendarWidget()
        self.end_date_edit = QCalendarWidget()

        self.result_browser = QTextBrowser()
        self.show_button = QPushButton("Показать сводку")
        self.export_button = QPushButton("Экспорт в Excel")
        self.close_button = QPushButton("Закрыть")

        self.show_button.clicked.connect(self.show_summary)
        self.export_button.clicked.connect(self.export_to_excel)
        self.close_button.clicked.connect(self.accept)

        layout = QVBoxLayout()
        layout.addWidget(QLabel("Начальная дата:"))
        layout.addWidget(self.start_date_edit)
        layout.addWidget(QLabel("Конечная дата:"))
        layout.addWidget(self.end_date_edit)
        layout.addWidget(self.show_button)
        layout.addWidget(self.export_button)
        layout.addWidget(self.result_browser)
        layout.addWidget(self.close_button)
        self.setLayout(layout)

    def show_summary(self):
        start_date = self.start_date_edit.selectedDate().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.selectedDate().toString("yyyy-MM-dd")

        result = self.db.fetch_all("""
            SELECT e.first_name, e.last_name, SUM(s.shift_salary) as total_salary, COUNT(*) as shift_count
            FROM shifts s
            JOIN employees e ON s.employee_id = e.id
            WHERE s.shift_date BETWEEN %s AND %s
            GROUP BY e.id, e.first_name, e.last_name
            ORDER BY total_salary DESC
        """, (start_date, end_date))

        self.summary_data = result
        self.result_browser.clear()
        self.result_browser.append(f"Сводка с {start_date} по {end_date}:\n")

        if not result:
            self.result_browser.append("Нет данных за выбранный период.")
            return

        for row in result:
            full_name = f"{row['first_name']} {row['last_name']}"
            self.result_browser.append(
                f"👤 {full_name} — 💰 {row['total_salary']:.2f} ₽ ({row['shift_count']} смен)"
            )

    def export_to_excel(self):
        start_date = self.start_date_edit.selectedDate().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.selectedDate().toString("yyyy-MM-dd")

        if not self.summary_data:
            QMessageBox.warning(self, "Нет данных", "Сначала сформируйте сводку.")
            return

        try:
            wb = Workbook()
            wb.remove(wb.active)

            for row in self.summary_data:
                emp_name = f"{row['first_name']} {row['last_name']}"
                ws = wb.create_sheet(title=emp_name[:31])

                ws.append(["Дата", "Начало", "Конец", "Часы", "Зарплата"])
                ws.row_dimensions[1].font = Font(bold=True)

                shifts = self.db.fetch_all("""
                    SELECT shift_date, shift_start, shift_end, shift_salary
                    FROM shifts
                    WHERE employee_id = (
                        SELECT id FROM employees WHERE first_name = %s AND last_name = %s
                    ) AND shift_date BETWEEN %s AND %s
                    ORDER BY shift_date
                """, (row['first_name'], row['last_name'], start_date, end_date))

                total_salary = 0
                total_hours = 0

                for shift in shifts:
                    start = to_qtime(shift['shift_start'])
                    end = to_qtime(shift['shift_end'])
                    hours = round(start.secsTo(end) / 3600, 2)
                    salary = float(shift['shift_salary'])

                    total_salary += salary
                    total_hours += hours

                    ws.append([
                        shift['shift_date'],
                        shift['shift_start'],
                        shift['shift_end'],
                        hours,
                        salary
                    ])

                ws.append([])
                ws.append(["Итого", "", "", total_hours, total_salary])
                ws.row_dimensions[ws.max_row].font = Font(bold=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            temp_path = os.path.join(tempfile.gettempdir(), f"Сводка_{timestamp}.xlsx")
            wb.save(temp_path)

            system = platform.system()
            if system == "Windows":
                os.startfile(temp_path)
            elif system == "Darwin":  # macO
                subprocess.call(["open", temp_path])
            elif system == "Linux":
                subprocess.call(["xdg-open", temp_path])
            else:
                QMessageBox.information(self, "Инфо", f"Файл сохранён: {temp_path}\nОткройте его вручную.")

            QMessageBox.information(self, "Готово", f"Файл экспортирован: {temp_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте:\n{e}")


class EmployeeTab(QWidget):
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)

        self.employee_selector = QComboBox()
        self.employee_selector.currentIndexChanged.connect(self.load_shifts)
        layout.addWidget(QLabel("Выберите сотрудника:"))
        layout.addWidget(self.employee_selector)

        self.calendar = QCalendarWidget()
        self.calendar.clicked.connect(self.on_calendar_clicked)
        layout.addWidget(self.calendar)

        self.info_browser = QTextBrowser()
        layout.addWidget(QLabel("Смены и зарплата:"))
        layout.addWidget(self.info_browser)

        self.manage_button = QPushButton("Управление сменами")
        self.manage_button.clicked.connect(self.open_manage_shifts)
        layout.addWidget(self.manage_button)

        self.summary_button = QPushButton("Сводка")
        self.summary_button.clicked.connect(self.open_summary)
        layout.addWidget(self.summary_button)

        self.refresh_button = QPushButton("Обновить данные")
        self.refresh_button.clicked.connect(self.load_employees)
        layout.addWidget(self.refresh_button)

        self.add_employee_button = QPushButton("Добавить сотрудника")
        self.add_employee_button.clicked.connect(self.add_employee)
        layout.addWidget(self.add_employee_button)

        self.load_employees()

    def load_employees(self):
        employees = self.db.fetch_all("SELECT id, first_name, last_name FROM employees")
        self.employee_selector.clear()
        self.employees_data = {}

        for emp in employees:
            full_name = f"{emp['first_name']} {emp['last_name']}"
            self.employee_selector.addItem(full_name, emp['id'])
            self.employees_data[emp['id']] = full_name

        if employees:
            self.load_shifts()

    def load_shifts(self):
        emp_id = self.employee_selector.currentData()
        if not emp_id:
            return

        shifts = self.db.fetch_all("""
            SELECT shift_date, shift_start, shift_end, shift_salary, employee_id
            FROM shifts
            WHERE employee_id = %s
            ORDER BY shift_date
        """, (emp_id,))

        self.info_browser.clear()
        self.info_browser.append(f"Смен в месяц: {len(shifts)}\n📅 Смены:")

        total = 0
        dates = []

        for shift in shifts:
            date = QDate.fromString(str(shift['shift_date']), "yyyy-MM-dd")
            dates.append(date)
            salary = shift['shift_salary'] or 0
            total += salary
            self.info_browser.append(
                f"— {date.toString('dd.MM.yyyy')} ({shift['shift_start']}–{shift['shift_end']}) — {salary:.2f} ₽"
            )

        self.info_browser.append(f"\n💰 Всего начислено: {total:.2f} ₽")
        self.highlight_calendar_days(dates)

    def highlight_calendar_days(self, dates):
        if dates:
            self.calendar.setSelectedDate(dates[-1])

    def on_calendar_clicked(self, date: QDate):
        emp_id = self.employee_selector.currentData()
        if not emp_id:
            QMessageBox.warning(self, "Ошибка", "Сначала выберите сотрудника.")
            return

        shift_date = date.toString("yyyy-MM-dd")

        try:
            shift = self.db.fetch_one(
                "SELECT * FROM shifts WHERE employee_id = %s AND shift_date = %s",
                (emp_id, shift_date)
            )

            if shift:
                reply = QMessageBox.question(self, "Смена существует", "Смена уже существует. Изменить?",
                                             QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                if reply == QMessageBox.StandardButton.Yes:
                    dialog = EditShiftDialog(self, shift_data=shift, employees=self.employees_data)
                    if dialog.exec():
                        self._save_shift(dialog.get_data(), shift_date, existing_id=shift['id'])
            else:
                dialog = EditShiftDialog(self, employees=self.employees_data)
                if dialog.exec():
                    data = dialog.get_data()
                    if self.db.fetch_one("SELECT * FROM shifts WHERE shift_date = %s AND employee_id = %s",
                                         (shift_date, data[0])):
                        QMessageBox.warning(self, "Ошибка", "Этот сотрудник уже работает в этот день.")
                        return
                    self._save_shift(data, shift_date)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при обработке смены:\n{e}")

    def _save_shift(self, data, shift_date, existing_id=None):
        employee_id, start_time, end_time, hourly_rate = data
        start = QTime.fromString(start_time, "HH:mm:ss")
        end = QTime.fromString(end_time, "HH:mm:ss")

        if not start.isValid() or not end.isValid() or start >= end:
            QMessageBox.warning(self, "Ошибка", "Некорректное время.")
            return

        hours = start.secsTo(end) / 3600
        total_salary = round(hourly_rate * hours, 2)

        if existing_id:
            self.db.execute("""
                UPDATE shifts
                SET shift_start = %s, shift_end = %s, shift_salary = %s, employee_id = %s
                WHERE id = %s
            """, (start_time, end_time, total_salary, employee_id, existing_id))
            QMessageBox.information(self, "Обновлено", "Смена обновлена.")
        else:
            self.db.execute("""
                INSERT INTO shifts (employee_id, shift_date, shift_start, shift_end, shift_salary)
                VALUES (%s, %s, %s, %s, %s)
            """, (employee_id, shift_date, start_time, end_time, total_salary))
            QMessageBox.information(self, "Добавлено", f"Смена на {shift_date} добавлена.")

        self.load_shifts()

    def open_manage_shifts(self):
        try:
            from ui.manage_shifts_dialog import ManageShiftsDialog
            dialog = ManageShiftsDialog(self, db=self.db, employees=self.employees_data)
            if dialog.exec():
                self.load_shifts()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть управление сменами:\n{e}")

    def open_summary(self):
        dialog = SummaryDialog(self, db=self.db, employees=self.employees_data)
        dialog.exec()

    def add_employee(self):
        dialog = AddEmployeeDialog(self)
        if dialog.exec():
            first_name, last_name = dialog.get_data()
            if not first_name or not last_name:
                QMessageBox.warning(self, "Ошибка", "Имя и фамилия обязательны.")
                return
            self.db.execute("""
                INSERT INTO employees (first_name, last_name)
                VALUES (%s, %s)
            """, (first_name, last_name))
            QMessageBox.information(self, "Добавлено", "Сотрудник успешно добавлен.")
            self.load_employees()
